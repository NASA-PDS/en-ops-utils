"""
Data Release Calendar (DRC) Update Script

This script manages the PDS Data Release Calendar webpage and associated files.
It handles Excel spreadsheet creation, HTML updates, JSON generation, and
year-end rollover tasks.

Usage:
    # Create/open Excel for today's date
    python drc.py --create-excel

    # Create Excel with specific version
    python drc.py --create-excel --excel-version 2

    # Update HTML and JSON after editing Excel
    python drc.py --update-files

    # Initialize files for new year (requires manual ROSES link cleanup)
    python drc.py --initialize-new --year 2027

    # Finalize new year setup (updates JS and redirect)
    python drc.py --finish-new --year 2027

Environment Variables:
    DRC_REPO_PATH: Path to portal-legacy repository (required)
"""

import argparse
import calendar
import json
import logging
import os
import re
import shutil
import subprocess
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional, List, Callable, Tuple
from openpyxl import load_workbook


# Logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(levelname)s: %(lineno)d: %(message)s'
)
logger = logging.getLogger(__name__)


# Configuration
REPO_PATH = Path(os.environ.get('DRC_REPO_PATH', '/path/to/default'))
if not REPO_PATH.exists():
    sys.exit(f'ERROR: Repository path not found: {REPO_PATH}. Set DRC_REPO_PATH environment variable.')

DRC_PATH = Path('datasearch/subscription-service')

# Global state (to be minimized over time)
DEBUG = False
indent = '  '


# ============================================================================
# CUSTOM EXCEPTIONS
# ============================================================================

class DRCError(Exception):
    """Base exception for DRC operations."""
    pass


class ExcelFileNotFoundError(DRCError):
    """Excel file not found at specified path."""
    pass


class InvalidWorksheetError(DRCError):
    """Worksheet not found in Excel file."""
    pass


class InvalidYearError(DRCError):
    """Invalid year format provided."""
    pass


class HTMLFileNotFoundError(DRCError):
    """No HTML files found in DRC directory."""
    pass


class InvalidHTMLFilenameError(DRCError):
    """Could not parse year from HTML filename."""
    pass


class YearOrderError(DRCError):
    """Requested year is not after the most recent year."""
    pass


class MissingDateCellError(DRCError):
    """Required date cell is blank in Excel."""
    pass


# ============================================================================
# DATA CLASS
# ============================================================================

@dataclass
class DRCExcelFile:
    """
    Represents a Data Release Calendar Excel file.

    Attributes:
        date: File date in YYYYMMDD format (e.g., '20260818')
        version: Version number (e.g., 3 for 'v3')
        path: Full path to the Excel file
    """
    date: str
    version: int
    path: Path

    # Class constant
    FILENAME_PREFIX = 'PDS_Data_Release_Calendar'

    @property
    def filename(self) -> str:
        """Generate standardized filename from metadata."""
        return f'{self.FILENAME_PREFIX}_{self.date}_v{self.version}.xlsx'

    @classmethod
    def get_filename_pattern(cls) -> str:
        """
        Get regex pattern for matching DRC Excel filenames.

        Returns:
            Regex pattern string

        Example:
            >>> pattern = DRCExcelFile.get_filename_pattern()
            >>> re.search(pattern, 'PDS_Data_Release_Calendar_20260818_v3.xlsx')
            <Match object>
        """
        return rf'{cls.FILENAME_PREFIX}_(\d{{8}})_v(\d+)\.xlsx'

    @classmethod
    def get_filename_glob(cls, date_prefix: str = '') -> str:
        """
        Get glob pattern for finding DRC Excel files.

        Args:
            date_prefix: Optional date prefix to filter (e.g., '202608')

        Returns:
            Glob pattern string

        Example:
            >>> DRCExcelFile.get_filename_glob('202608')
            'PDS_Data_Release_Calendar_202608*.xlsx'
        """
        return f'{cls.FILENAME_PREFIX}_{date_prefix}*.xlsx'

    @classmethod
    def get_html_link_pattern(cls, drc_path: str) -> str:
        """
        Get regex pattern for matching HTML links to Excel files.

        Args:
            drc_path: DRC path for URL (e.g., 'datasearch/subscription-service')

        Returns:
            Regex pattern for matching <a> tags linking to Excel files
        """
        return f'<a href="/?{drc_path}/{cls.FILENAME_PREFIX}_\\d{{8}}_v\\d{{1,2}}\\.xlsx">'

    def get_html_link(self, drc_path: str) -> str:
        """
        Generate HTML link tag for this file.

        Args:
            drc_path: DRC path for URL

        Returns:
            HTML <a> tag opening for this file
        """
        return f'<a href="/{drc_path}/{self.filename}">'

    @classmethod
    def from_path(cls, path: Path) -> Optional['DRCExcelFile']:
        """
        Parse an Excel file path into structured metadata.

        Args:
            path: Path to Excel file

        Returns:
            DRCExcelFile instance if path matches expected format, None otherwise

        Example:
            >>> path = Path('PDS_Data_Release_Calendar_20260818_v3.xlsx')
            >>> excel = DRCExcelFile.from_path(path)
            >>> excel.date
            '20260818'
            >>> excel.version
            3
        """
        pattern = cls.get_filename_pattern()
        match = re.search(pattern, path.name)
        if match:
            return cls(
                date=match.group(1),
                version=int(match.group(2)),
                path=path
            )
        return None

    @classmethod
    def create(cls, date: str, version: int, base_dir: Path) -> 'DRCExcelFile':
        """
        Create a new DRCExcelFile instance for a given date and version.

        Args:
            date: Date in YYYYMMDD format
            version: Version number
            base_dir: Directory where file should be located

        Returns:
            New DRCExcelFile instance
        """
        filename = f'{cls.FILENAME_PREFIX}_{date}_v{version}.xlsx'
        return cls(date=date, version=version, path=base_dir / filename)

    def is_same_date(self, other: 'DRCExcelFile') -> bool:
        """
        Check if this file has the same date as another file.

        Args:
            other: Another DRCExcelFile to compare

        Returns:
            True if dates match, False otherwise
        """
        return self.date == other.date

    def validate(self) -> None:
        """
        Validate file metadata.

        Raises:
            ValueError: If date or version format is invalid
        """
        if len(self.date) != 8 or not self.date.isdigit():
            raise ValueError(f"Invalid date format: {self.date}. Expected YYYYMMDD.")
        if self.version < 0:
            raise ValueError(f"Invalid version: {self.version}. Must be non-negative.")

    def __lt__(self, other: 'DRCExcelFile') -> bool:
        """Enable sorting by date (primary), then version (secondary)."""
        return (self.date, self.version) < (other.date, other.version)

    def __str__(self) -> str:
        """Human-readable representation."""
        return f'{self.filename} ({self.date} v{self.version})'


# ============================================================================
# FILE UTILITIES
# ============================================================================

def update_file_with_regex(file_path: Path, replacements: List[Tuple[str, str]]) -> None:
    """
    Update file by applying multiple regex replacements in sequence.

    This eliminates the repeated pattern of opening a file, reading content,
    applying regex substitutions, and writing back.

    Args:
        file_path: Path to file to update
        replacements: List of (pattern, replacement) tuples to apply in order

    Example:
        >>> replacements = [
        ...     (r'old_pattern', 'new_value'),
        ...     (r'another_pattern', 'another_value')
        ... ]
        >>> update_file_with_regex(path, replacements)
    """
    with open(file_path, 'r+') as f:
        content = f.read()
        for pattern, replacement in replacements:
            content = re.sub(pattern, replacement, content)
        f.seek(0)
        f.write(content)
        f.truncate()


# ============================================================================
# SERVICE CLASS FOR BUSINESS LOGIC
# ============================================================================

class DRCService:
    """
    Service class encapsulating DRC business logic.

    This separates business logic from I/O operations and improves testability.
    """

    def __init__(self, drc_dir: Path):
        """
        Initialize DRC service.

        Args:
            drc_dir: Path to DRC directory (datasearch/subscription-service)
        """
        self.drc_dir = drc_dir

    def find_latest_excel(self, up_to_date: str) -> Optional[DRCExcelFile]:
        """
        Find the most recent Excel file up to and including the specified date.

        Args:
            up_to_date: Maximum date in YYYYMMDD format (e.g., '20260819').
                       Files dated after this are excluded (e.g., future planning entries).

        Returns:
            Most recent DRCExcelFile with date <= up_to_date, or None if no files found

        Note:
            Searches all Excel files but filters to only those dated on or before
            up_to_date. This excludes future-dated planning entries while finding
            the absolute latest file as of the specified date.

        Example:
            >>> # Today is 2026-08-19
            >>> # Files exist: 20260715_v1, 20260820_v2, 20260920_v1 (future)
            >>> service.find_latest_excel('20260819')
            DRCExcelFile(date='20260715', version=1)  # Excludes future dates
        """
        excel_files = []
        pattern = DRCExcelFile.get_filename_glob()

        for path in self.drc_dir.glob(pattern):
            excel_file = DRCExcelFile.from_path(path)
            if excel_file and excel_file.date <= up_to_date:
                excel_files.append(excel_file)

        if excel_files:
            return max(excel_files)  # Uses __lt__ to sort by date, version

        return None

    def determine_target_excel(
        self,
        latest: DRCExcelFile,
        target_date: str,
        requested_version: Optional[int],
        allow_interactive: bool
    ) -> DRCExcelFile:
        """
        Determine which Excel file to create or open.

        Args:
            latest: Most recent existing Excel file
            target_date: Desired date in YYYYMMDD format
            requested_version: User-requested version number (None for auto-increment)
            allow_interactive: Whether to prompt user for decisions

        Returns:
            DRCExcelFile representing the file to use/create

        Logic:
            - Different date → start at version 0
            - Same date, no version specified → auto-increment latest version
            - Same date, version ≤ latest → prompt user or auto-increment
            - Same date, version > latest → use requested version
        """
        # Different date? Start fresh at version 0
        if not latest.is_same_date(DRCExcelFile.create(target_date, 0, self.drc_dir)):
            return DRCExcelFile.create(target_date, 0, self.drc_dir)

        # Same date - determine version
        if requested_version is None:
            # No preference - auto increment
            next_version = latest.version + 1
            logger.info('Auto-incrementing version: %d → %d', latest.version, next_version)
        elif requested_version <= latest.version:
            # Requested version already exists
            logger.info('Excel for %s v%d already exists.', target_date, requested_version)

            if allow_interactive:
                next_version = self._prompt_user_for_version(latest.version, requested_version)
            else:
                logger.info('Auto-incrementing version.')
                next_version = latest.version + 1
        else:
            # Requested version is higher than latest
            next_version = requested_version

        return DRCExcelFile.create(target_date, next_version, self.drc_dir)

    def _prompt_user_for_version(self, latest_version: int, requested_version: int) -> int:
        """
        Interactively ask user whether to use existing version or increment.

        Args:
            latest_version: Most recent version number
            requested_version: Version user initially requested

        Returns:
            Version number selected by user
        """
        prompt = ''
        if latest_version > requested_version:
            prompt = (f'There\'s a later version (v{latest_version}) and no foreseeable reason '
                     f'to edit the older version. Assuming you want the latest. ')

        while True:
            response = input(prompt + 'Proceed WITHOUT incrementing the version? (y/n) ')
            if response.lower() in ['y', 'yes']:
                print(indent, 'Okay. Using the existing version and file.')
                return latest_version
            elif response.lower() in ['n', 'no']:
                print(indent, 'Okay. Incrementing the version.')
                return latest_version + 1
            else:
                print(indent, 'Invalid response. Please enter y or n.')

    def create_or_open_excel(self, target: DRCExcelFile, latest: DRCExcelFile,
                            callback: Callable) -> None:
        """
        Create new Excel file (from latest) or open existing, then execute callback.

        Args:
            target: Excel file to create/open
            latest: Most recent Excel file (source for copying)
            callback: Function to call after opening file (e.g., open_changelog)

        Side Effects:
            - Creates new file by copying if target doesn't exist
            - Opens file in default application
            - Executes callback function
        """
        if target.path.exists():
            action = "Opened existing"
        else:
            shutil.copyfile(latest.path, target.path)
            action = "Created and opened new"

        subprocess.run(['open', str(target.path)], check=True)
        logger.info('%s Excel: %s', action, target.filename)

        callback()


# ============================================================================
# ARGUMENT PARSING
# ============================================================================

def parse_arguments():
    """
    Parse command-line arguments.

    Returns:
        Dictionary of parsed arguments
    """
    parser = argparse.ArgumentParser(
        description='Manage PDS Data Release Calendar files',
        epilog='Set DRC_REPO_PATH environment variable to portal-legacy repo path'
    )

    parser.add_argument('-c', '--create-excel',
        help='Create Excel with today\'s date from the most recent spreadsheet.',
        action='store_true')
    parser.add_argument('-e', '--excel-version',
        help='Use with `-c, --create-excel`. Specifies version number in the Excel filename.',
        type=int,
        metavar='')

    parser.add_argument('-u', '--update-files',
        help='Update the HTML and JSON files after editing Excel.',
        action='store_true')

    parser.add_argument('-i', '--initialize-new',
        help='Initialize files for a new year. Requires `--year YEAR`. Manual ROSES link cleanup still needed.',
        action='store_true')
    parser.add_argument('-f', '--finish-new',
        help='Update JS and redirect HTML for a new year. Requires `--year YEAR`.',
        action='store_true')

    parser.add_argument('-y', '--year',
        help='Specific year (YYYY format) for `--initialize-new` or `--finish-new`.',
        type=int,
        metavar='')

    parser.add_argument('--debug',
        help='Enable debug logging',
        action='store_true')

    return vars(parser.parse_args())


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def get_drc_dir() -> Path:
    """
    Get the full path to the DRC directory.

    Returns:
        Path to datasearch/subscription-service directory
    """
    return REPO_PATH / DRC_PATH


def open_changelog():
    """Open the changelog file for user to document changes."""
    changelog_path = get_drc_dir() / 'changelog.txt'
    subprocess.run(['open', str(changelog_path)], check=True)
    logger.info('Opening changelog.txt')


def get_latest_html_path() -> Path:
    """
    Find the most recent HTML file.

    Returns:
        Path to most recent data-release-calendar-YYYY.shtml file

    Raises:
        HTMLFileNotFoundError: If no HTML files found
    """
    drc_dir = get_drc_dir()
    htmls = sorted(drc_dir.glob('data-release-calendar-*.shtml'))
    if not htmls:
        raise HTMLFileNotFoundError('No HTML files found in DRC directory')
    return htmls[-1]


def get_year_from_html_filename(html_path: Path) -> str:
    """
    Extract year from HTML filename.

    Args:
        html_path: Path to HTML file

    Returns:
        Four-digit year string

    Raises:
        InvalidHTMLFilenameError: If year cannot be parsed from filename

    Example:
        >>> path = Path('data-release-calendar-2026.shtml')
        >>> get_year_from_html_filename(path)
        '2026'
    """
    match = re.search(r'data-release-calendar-(\d{4})\.shtml', html_path.name)
    if not match:
        raise InvalidHTMLFilenameError(f'Could not parse year from filename: {html_path.name}')
    return match.group(1)


def validate_year_string(year: str) -> None:
    """
    Validate year string format.

    Args:
        year: Year string to validate

    Raises:
        InvalidYearError: If year format is invalid
    """
    if not (len(year) == 4 and year.isdigit()):
        raise InvalidYearError(f'Invalid year format: {year}. Must be 4-digit year (e.g., 2027).')


# ============================================================================
# HTML OPERATIONS
# ============================================================================

def initialize_new_html_for(year: str):
    """
    Create a new HTML file for the specified year based on most recent year.

    Args:
        year: Four-digit year string (e.g., '2027')

    Side Effects:
        - Copies most recent HTML file
        - Replaces year references (except ROSES links)
        - Adds placeholder text for new year's ROSES links

    Raises:
        YearOrderError: If requested year is not after the most recent year
    """
    logger.info('Initializing new HTML for year %s', year)

    latest_html_path = get_latest_html_path()
    latest_year = get_year_from_html_filename(latest_html_path)

    logger.debug('Latest HTML path: %s', latest_html_path)
    logger.debug('Latest year: %s', latest_year)

    if int(year) <= int(latest_year):
        raise YearOrderError(
            f'Requested year {year} must be after most recent year {latest_year}.'
        )

    new_html_path = get_drc_dir() / f'data-release-calendar-{year}.shtml'

    # Copy and modify HTML file
    subprocess.run(['cp', str(latest_html_path), str(new_html_path)], check=True)

    with open(latest_html_path, 'r') as reader:
        with open(new_html_path, 'w') as writer:
            for line in reader:
                if latest_year in line:
                    if "ROSES" in line:
                        # Keep the old ROSES link
                        writer.write(line)
                        # Add placeholder for new year's ROSES
                        tabs = '\t' * 5
                        writer.write(f'{tabs}<br/>\n')
                        writer.write(f'{tabs}Due dates for {year} ROSES program elements '
                                   f'will be linked here when available.\n')
                    else:
                        # Replace year in non-ROSES lines
                        new_line = line.replace(latest_year, year)
                        writer.write(new_line)
                else:
                    writer.write(line)

    logger.info('Created new HTML for %s', year)


def update_html_for(year: str, latest_excel: DRCExcelFile, current_date: date,
                   initialize_new: bool = False):
    """
    Update HTML file with current Excel filename and last updated date.

    Args:
        year: Four-digit year string
        latest_excel: Most recent Excel file (after any edits)
        current_date: Today's date for "Last updated" text
        initialize_new: Whether to initialize a new HTML file first

    Side Effects:
        - Modifies HTML file in place
        - Updates Excel download link
        - Updates "Last updated on" date string
    """
    if initialize_new:
        initialize_new_html_for(year)

    html_path = get_drc_dir() / f'data-release-calendar-{year}.shtml'
    drc_path_str = str(DRC_PATH).replace('\\', '/')  # Ensure forward slashes for URLs

    # Use class methods for pattern generation
    link_pattern = DRCExcelFile.get_html_link_pattern(drc_path_str)
    link_replacement = latest_excel.get_html_link(drc_path_str)

    date_pattern = r'Last updated on \w+ \d{2}, \d{4}\.'
    month_name = calendar.month_name[current_date.month]
    date_replacement = f'Last updated on {month_name} {current_date.day:02d}, {current_date.year}.'

    # Use helper function for file update
    update_file_with_regex(html_path, [
        (link_pattern, link_replacement),
        (date_pattern, date_replacement)
    ])

    logger.info('Updated HTML for year %s', year)


# ============================================================================
# EXCEL OPERATIONS
# ============================================================================

def massage_primary_target_column(cell_value) -> str:
    """
    Clean primary target cell value by removing problematic characters.

    Args:
        cell_value: Raw cell value from Excel

    Returns:
        Cleaned string with newlines, slashes, hyphens, and commas removed
    """
    if cell_value is None:
        return ''
    return str(cell_value).replace('\n', ' ').replace('/', ' ').replace('-', ' ').replace(',', '')


def massage_datetime_column(cell_value: datetime) -> str:
    """
    Format datetime cell as MM/DD/YYYY string.

    Args:
        cell_value: datetime object from Excel

    Returns:
        Formatted date string
    """
    return cell_value.strftime('%m/%d/%Y')


def massage_link_column(cell):
    """
    Extract hyperlink URL from Excel cell.

    Args:
        cell: openpyxl Cell object with hyperlink

    Returns:
        URL string from hyperlink target
    """
    return cell.hyperlink.target


def read_excel_file_for(year: str, excel_file: DRCExcelFile, callback: Callable):
    """
    Parse Excel file and extract data rows, then pass to callback.

    Args:
        year: Four-digit year (worksheet name)
        excel_file: Excel file to read
        callback: Function to call with (year, data_list)

    Data Structure:
        Each row is a list with 8 elements:
        [0] Column B: Mission
        [1] Column C: Instrument
        [2] Column D: Release Number/Name
        [3] Column E: Release Interval
        [4] Column F: Estimated Release Date
        [5] Column G: Actual Release Date
        [6] Column H: Link
        [7] Column A: Primary Target

    Raises:
        ExcelFileNotFoundError: If Excel file not found
        InvalidWorksheetError: If worksheet not found
        MissingDateCellError: If date cells are empty
    """
    logger.debug('Reading Excel file: %s, worksheet: %s', excel_file.filename, year)

    # Load workbook
    try:
        wb = load_workbook(excel_file.path)
    except FileNotFoundError:
        raise ExcelFileNotFoundError(f'Excel file not found: {excel_file.path}')
    except Exception as e:
        raise DRCError(f'Failed to load Excel file: {e}')

    # Get worksheet
    try:
        ws = wb[year]
    except KeyError:
        available = ', '.join(wb.sheetnames)
        raise InvalidWorksheetError(
            f'Worksheet "{year}" not found in {excel_file.filename}. '
            f'Available worksheets: {available}'
        )

    # Parse rows (skip header row 1)
    data_list = []
    for i in range(2, ws.max_row + 1):
        logger.debug('Processing row %d', i)

        data_row = [None] * 8

        for j, cell in enumerate(ws[i]):
            logger.debug('  Column %d', j)

            if j == 0:  # Column A: Primary Target → data_row[7]
                data_row[7] = massage_primary_target_column(cell.value)

            elif j == 3:  # Column D: Release Number/Name (may have newlines)
                data_row[j - 1] = str(cell.value).replace('\n', ' ')

            elif j == 5 or j == 6:  # Columns F-G: Release dates
                if cell.value is None:
                    raise MissingDateCellError(
                        f'Cell value is blank for DATE in row {i} of {excel_file.filename}. '
                        f'Column {chr(65 + j)} must contain a date.'
                    )

                if isinstance(cell.value, datetime):
                    value = massage_datetime_column(cell.value)
                else:
                    value = cell.value
                data_row[j - 1] = value

            elif j == 7:  # Column H: Link (extract hyperlink)
                if cell.value == 'Link':
                    value = massage_link_column(cell)
                else:
                    value = cell.value
                data_row[j - 1] = value

            elif j > 7:  # Ignore columns beyond H
                logger.debug('  Ignoring column %d (beyond expected range)', j)
                break

            else:  # Columns B, C, E (Mission, Instrument, Release Interval)
                data_row[j - 1] = cell.value

        data_list.append(data_row)

    callback(year, data_list)


# ============================================================================
# JSON OPERATIONS
# ============================================================================

def update_json_for(year: str, excel_file: DRCExcelFile):
    """
    Generate JSON file from Excel data.

    Args:
        year: Four-digit year (worksheet name)
        excel_file: Excel file to read

    Side Effects:
        Creates/updates data-release-calendar-YYYY.txt with JSON data
    """
    def write_json(year: str, data_list: List):
        """Callback to write parsed data as JSON."""
        logger.debug('Writing TXT with JSON for year %s', year)

        json_path = get_drc_dir() / f'data-release-calendar-{year}.txt'
        excel_dict = {'data': data_list}

        with open(json_path, 'w') as f:
            json.dump(excel_dict, f, indent=4)

        logger.info('Updated TXT with JSON for year %s', year)

    read_excel_file_for(year, excel_file, write_json)


# ============================================================================
# COMPOSITE OPERATIONS
# ============================================================================

def update_regulars_for(year: str, latest_excel: DRCExcelFile, current_date: date,
                       initialize_new: bool = False):
    """
    Update the "regular" files: TXT (JSON) and HTML.

    Args:
        year: Four-digit year string
        latest_excel: Most recent Excel file
        current_date: Today's date
        initialize_new: Whether to initialize new HTML first

    Note:
        Called after Excel has been edited with new data.
    """
    update_json_for(year, latest_excel)
    update_html_for(year, latest_excel, current_date, initialize_new)


def update_support_for(year: str):
    """
    Update the "support" files for a new year: JS and redirect HTML.

    Args:
        year: Four-digit year string

    Side Effects:
        - Updates js/drc.js with new end_year value
        - Updates data-release-calendar.shtml redirect to new year

    Note:
        Only called after a new year's HTML page has been initialized.
    """
    logger.info('Updating support files (JS and redirect) for year %s', year)

    # Update JavaScript file
    js_file = REPO_PATH / 'js' / 'drc.js'
    js_replacements = [(r', end_year = \d{4}', f', end_year = {year}')]
    update_file_with_regex(js_file, js_replacements)

    # Update redirect HTML
    redirect_file = get_drc_dir() / 'data-release-calendar.shtml'
    redirect_replacements = [
        (r'<meta http-equiv="refresh" content="0; '
         r'url=/datasearch/subscription-service/data-release-calendar-\d{4}\.shtml">',
         f'<meta http-equiv="refresh" content="0; '
         f'url=/datasearch/subscription-service/data-release-calendar-{year}.shtml">')
    ]
    update_file_with_regex(redirect_file, redirect_replacements)


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main(**args):
    """
    Main entry point for DRC script.

    Args:
        **args: Parsed command-line arguments from parse_arguments()

    Workflow:
        1. Initialize service with DRC directory
        2. Find most recent Excel file
        Do one of the following four:
           - Handle create-excel: determine target, create/open, launch changelog
           - Handle update-files: generate JSON and update HTML
           - Handle initialize-new: create new year's HTML/JSON
           - Handle finish-new: update JS and redirect for new year
    """

    if args['debug']:
        logger.setLevel(logging.DEBUG)
        logger.debug('Command-line arguments:')
        for k, v in args.items():
            logger.debug('%s%s: %s', indent, k, v)

    try:
        # Initialize service
        service = DRCService(get_drc_dir())

        # Get today's date
        today = date.today()
        target_date = today.strftime('%Y%m%d')

        # Find most recent Excel file up to today (excludes future-dated planning entries)
        latest_excel = service.find_latest_excel(target_date)
        if not latest_excel:
            raise ExcelFileNotFoundError('No existing Excel files found up to today. Cannot proceed.')

        # Handle create-excel command
        if args['create_excel']:
            requested_version = args.get('excel_version')
            allow_interactive = requested_version is not None

            target_excel = service.determine_target_excel(
                latest_excel,
                target_date,
                requested_version,
                allow_interactive
            )

            service.create_or_open_excel(target_excel, latest_excel, open_changelog)

            # Update latest_excel reference for potential subsequent operations
            latest_excel = target_excel

        # Handle update-files command
        elif args['update_files']:
            year = today.strftime('%Y')
            update_regulars_for(year, latest_excel, today, initialize_new=False)

        # Handle year-based commands
        elif args['year']:
            year = str(args['year'])

            # Validate year format
            validate_year_string(year)

            if args['initialize_new']:
                # Initialize new year's files
                update_regulars_for(year, latest_excel, today, initialize_new=True)

            elif args['finish_new']:
                # Finalize new year setup
                update_support_for(year)

            else:
                raise DRCError('Must specify `--initialize-new` or `--finish-new` with `--year`.')

        else:
            raise DRCError('Unrecognized combination of arguments. Use `--help` for usage information.')

    except DRCError as e:
        logger.error(str(e))
        sys.exit(1)
    except Exception as e:
        logger.exception('Unexpected error occurred')
        sys.exit(1)


if __name__ == '__main__':
    arguments = parse_arguments()
    main(**arguments)
